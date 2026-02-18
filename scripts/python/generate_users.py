#!/usr/bin/env python3
"""
Enterprise IAM - User Generation Script
Generates simulated user data for IAM testing and training

Author: Sheniah Ferguson
Purpose: Portfolio demonstration of Python automation for IAM
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import random
import string

# Company Configuration (Anonymized)
COMPANY = "Example Healthcare"
DOMAIN = "example.com"

# Sites
SITES = ["Site-A", "Site-B", "Site-C"]

# Departments and Roles for Healthcare
DEPARTMENTS = {
    "Clinical Operations": [
        "Registered Nurse",
        "Nurse Practitioner",
        "Medical Assistant",
        "Clinical Manager",
        "Care Coordinator"
    ],
    "Administration": [
        "Administrative Assistant",
        "Office Manager",
        "Executive Assistant",
        "Receptionist",
        "Scheduling Coordinator"
    ],
    "Finance": [
        "Billing Specialist",
        "Accounts Receivable Clerk",
        "Financial Analyst",
        "Revenue Cycle Manager",
        "Payroll Specialist"
    ],
    "IT": [
        "IT Support Specialist",
        "Systems Administrator",
        "Network Engineer",
        "IT Manager",
        "Help Desk Technician"
    ],
    "Human Resources": [
        "HR Generalist",
        "Recruiter",
        "HR Manager",
        "Benefits Coordinator",
        "Training Specialist"
    ],
    "Compliance": [
        "Compliance Officer",
        "HIPAA Analyst",
        "Quality Assurance Specialist",
        "Risk Manager",
        "Audit Coordinator"
    ],
    "Patient Services": [
        "Patient Access Representative",
        "Patient Advocate",
        "Medical Records Clerk",
        "Insurance Verification Specialist",
        "Referral Coordinator"
    ]
}

# M365 License Types
LICENSE_TYPES = [
    "Microsoft 365 E3",
    "Microsoft 365 E5",
    "Microsoft 365 Business Basic",
    "Microsoft 365 Business Standard",
    "Microsoft 365 F3"
]

# Name pools for random generation
FIRST_NAMES = [
    "James", "Mary", "Robert", "Patricia", "John", "Jennifer", "Michael", "Linda",
    "David", "Elizabeth", "William", "Barbara", "Richard", "Susan", "Joseph", "Jessica",
    "Thomas", "Sarah", "Christopher", "Karen", "Charles", "Lisa", "Daniel", "Nancy",
    "Matthew", "Betty", "Anthony", "Margaret", "Mark", "Sandra", "Donald", "Ashley"
]

LAST_NAMES = [
    "Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis",
    "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez", "Wilson", "Anderson", "Thomas",
    "Taylor", "Moore", "Jackson", "Martin", "Lee", "Perez", "Thompson", "White"
]


def generate_password(length=12):
    """Generate a secure random password meeting complexity requirements"""
    chars = string.ascii_letters + string.digits + "!@#$%"
    password = [
        random.choice(string.ascii_uppercase),
        random.choice(string.ascii_lowercase),
        random.choice(string.digits),
        random.choice("!@#$%")
    ]
    password += [random.choice(chars) for _ in range(length - 4)]
    random.shuffle(password)
    return ''.join(password)


def generate_employee_id():
    """Generate unique employee ID"""
    return f"EMP{random.randint(10000, 99999)}"


def generate_username(first_name, last_name, existing_usernames):
    """Generate unique username following naming convention"""
    base_username = f"{first_name[0].lower()}{last_name.lower()}"
    username = base_username
    counter = 1
    while username in existing_usernames:
        username = f"{base_username}{counter}"
        counter += 1
    return username


def determine_license(department, job_title):
    """Determine appropriate M365 license based on role"""
    if department == "IT" or "Manager" in job_title:
        return "Microsoft 365 E5"
    elif department in ["Finance", "Compliance"]:
        return "Microsoft 365 E3"
    elif department == "Administration":
        return "Microsoft 365 Business Standard"
    else:
        return random.choice(["Microsoft 365 E3", "Microsoft 365 Business Standard"])


def determine_azure_role(job_title):
    """Determine Azure AD role based on job title"""
    if "Manager" in job_title or "Director" in job_title:
        return "User Administrator"
    elif "Administrator" in job_title or "Officer" in job_title:
        return "Privileged Role Administrator"
    else:
        return "User"


def generate_users(num_users=100):
    """
    Generate simulated user data for IAM system

    Args:
        num_users: Number of users to generate

    Returns:
        List of user dictionaries
    """
    users = []
    existing_usernames = set()

    for i in range(num_users):
        first_name = random.choice(FIRST_NAMES)
        last_name = random.choice(LAST_NAMES)
        username = generate_username(first_name, last_name, existing_usernames)
        existing_usernames.add(username)

        department = random.choice(list(DEPARTMENTS.keys()))
        job_title = random.choice(DEPARTMENTS[department])
        site = random.choice(SITES)

        user = {
            "Employee_ID": generate_employee_id(),
            "First_Name": first_name,
            "Last_Name": last_name,
            "Display_Name": f"{first_name} {last_name}",
            "Username": username,
            "Email": f"{username}@{DOMAIN}",
            "Department": department,
            "Job_Title": job_title,
            "Site": site,
            "M365_License": determine_license(department, job_title),
            "Azure_AD_Role": determine_azure_role(job_title),
            "Account_Status": random.choices(
                ["Active", "Inactive"],
                weights=[95, 5]
            )[0],
            "MFA_Enabled": random.choices(
                ["Yes", "No"],
                weights=[90, 10]
            )[0],
        }
        users.append(user)

    return users


def create_excel_export(users, output_file):
    """
    Create formatted Excel workbook with multiple sheets

    Args:
        users: List of user dictionaries
        output_file: Path to save Excel file
    """
    wb = Workbook()

    # Sheet 1: All Users
    ws_users = wb.active
    ws_users.title = "All_Users"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    # Write headers
    headers = list(users[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws_users.cell(row=1, column=col, value=header.replace("_", " "))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Write data
    for row_num, user in enumerate(users, 2):
        for col_num, key in enumerate(headers, 1):
            ws_users.cell(row=row_num, column=col_num, value=user[key])

    # Sheet 2: Department Summary
    ws_dept = wb.create_sheet("Department_Summary")
    dept_counts = {}
    for user in users:
        dept = user["Department"]
        dept_counts[dept] = dept_counts.get(dept, 0) + 1

    ws_dept.cell(row=1, column=1, value="Department").font = header_font
    ws_dept.cell(row=1, column=1).fill = header_fill
    ws_dept.cell(row=1, column=2, value="User Count").font = header_font
    ws_dept.cell(row=1, column=2).fill = header_fill

    for row, (dept, count) in enumerate(sorted(dept_counts.items()), 2):
        ws_dept.cell(row=row, column=1, value=dept)
        ws_dept.cell(row=row, column=2, value=count)

    # Sheet 3: License Summary
    ws_lic = wb.create_sheet("License_Summary")
    lic_counts = {}
    for user in users:
        lic = user["M365_License"]
        lic_counts[lic] = lic_counts.get(lic, 0) + 1

    ws_lic.cell(row=1, column=1, value="License Type").font = header_font
    ws_lic.cell(row=1, column=1).fill = header_fill
    ws_lic.cell(row=1, column=2, value="Count").font = header_font
    ws_lic.cell(row=1, column=2).fill = header_fill

    for row, (lic, count) in enumerate(sorted(lic_counts.items()), 2):
        ws_lic.cell(row=row, column=1, value=lic)
        ws_lic.cell(row=row, column=2, value=count)

    wb.save(output_file)
    return output_file


def main():
    """Main execution function"""
    print("=" * 60)
    print("ENTERPRISE IAM - USER GENERATION SCRIPT")
    print("=" * 60)

    print("\nGenerating 100 users...")
    users = generate_users(100)

    print("Creating Excel export...")
    output_file = "IAM_Users_Export.xlsx"
    create_excel_export(users, output_file)

    print(f"\nExport saved to: {output_file}")
    print("\nSummary:")
    print(f"  Total Users: {len(users)}")
    print(f"  Departments: {len(DEPARTMENTS)}")
    print(f"  Sites: {len(SITES)}")

    # Print department breakdown
    dept_counts = {}
    for user in users:
        dept = user["Department"]
        dept_counts[dept] = dept_counts.get(dept, 0) + 1

    print("\nUsers by Department:")
    for dept, count in sorted(dept_counts.items()):
        print(f"  {dept}: {count}")

    print("\n" + "=" * 60)


if __name__ == "__main__":
    main()
